import json
import argparse
import openpyxl
from fpdf import FPDF

class RiskAnalysis:
    def __init__(self, project_name):
        self.project_name = project_name
        self.risks = []

    def identify_risk(self, description, likelihood, impact):
        risk = {
            'description': description,
            'likelihood': likelihood,
            'impact': impact,
            'risk_score': likelihood * impact
        }
        self.risks.append(risk)

    def evaluate_risks(self):
        print("Risk Evaluation:")
        for risk in self.risks:
            print(f"Risk: {risk['description']}")
            print(f"   Likelihood: {risk['likelihood']}")
            print(f"   Impact: {risk['impact']}")
            print(f"   Risk Score: {risk['risk_score']}")
            print()

    def plan_responses(self):
        print("Risk Responses:")
        for risk in self.risks:
            if risk['risk_score'] >= 10:
                print(f"Develop response plan for high-risk item: {risk['description']}")
            elif 5 <= risk['risk_score'] < 10:
                print(f"Monitor and manage: {risk['description']}")
            else:
                print(f"Acceptable risk: {risk['description']}")
        print()

    def export_to_pdf(self, filename):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        pdf.cell(200, 10, f"Rapport d'analyse de risque pour le projet {self.project_name}", 1, 1, "C")

        for risk in self.risks:
            pdf.cell(200, 10, f"Risk: {risk['description']}", 1, 1, "L")
            pdf.cell(200, 10, f"   Likelihood: {risk['likelihood']}", 1, 1, "L")
            pdf.cell(200, 10, f"   Impact: {risk['impact']}", 1, 1, "L")
            pdf.cell(200, 10, f"   Risk Score: {risk['risk_score']}", 1, 1, "L")
            pdf.ln()

        pdf.output(filename)

    def export_to_excel(self, filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Rapport de vulnérabilités"

        sheet.append(["Risk", "Likelihood", "Impact", "Risk Score"])
        for risk in self.risks:
            sheet.append([risk['description'], risk['likelihood'], risk['impact'], risk['risk_score']])

        workbook.save(filename)

    def export_to_json(self, filename):
        with open(filename, 'w') as json_file:
            json.dump({'project_name': self.project_name, 'risks': self.risks}, json_file)

# Exemple d'utilisation avec des arguments
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Script d'analyse de risque avec génération de rapports.")
    parser.add_argument("-t", "--type", choices=["pdf", "excel"], help="Type de rapport à générer (pdf ou excel)", required=True)
    args = parser.parse_args()

    project_name = "Projet XYZ"
    risk_analysis = RiskAnalysis(project_name)

    risk_analysis.identify_risk("Défaillance du serveur", likelihood=3, impact=4)
    risk_analysis.identify_risk("Retard dans le développement", likelihood=2, impact=5)
    risk_analysis.identify_risk("Changements dans les exigences du client", likelihood=4, impact=3)

    risk_analysis.evaluate_risks()
    risk_analysis.plan_responses()

    # Génère le rapport spécifié par l'utilisateur
    if args.type == "pdf":
        risk_analysis.export_to_pdf("rapport_analyse_risque.pdf")
    elif args.type == "excel":
        risk_analysis.export_to_excel("rapport_analyse_risque.xlsx")
    else:
        print("Type de rapport non pris en charge.")
